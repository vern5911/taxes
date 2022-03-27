# import openpyxl and tkinter modules
# source: https://tutorialspoint.dev/language/python/python-simple-registration-form-using-tkinter
import openpyxl
from openpyxl import load_workbook
from tkinter import *
import datetime

# globally declare wb and sheet variable 
  
# opening the existing spreadsheet file
wb = load_workbook('C:/Taxes/taxes-2021.xlsx') 
  
# create the sheet object 
ws = wb.active 

# Function to resize all spreadsheet columns and
# insure each has a header.  
def reset(): 
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10
    ws.column_dimensions['I'].width = 10
    # Now insert column headers 
    ws.cell(row=1, column=1).value = "Date"
    ws.cell(row=1, column=2).value = "Type"
    ws.cell(row=1, column=3).value = "Payee"
    ws.cell(row=1, column=4).value = "Charity"
    ws.cell(row=1, column=5).value = "Healthcare"
    ws.cell(row=1, column=6).value = "Auto"
    ws.cell(row=1, column=7).value = "Medical"
    ws.cell(row=1, column=8).value = "Taxes"
    ws.cell(row=1, column=9).value = "Other"
  
# set focus on the tdate_field box (cursor) 
def focus1(event): 
    tdate_field.focus_set() 
  
# set focus on the ttype_field box  
def focus2(event): 
    ttype_field.focus_set() 
  
# set focus on the tpayee_field box 
def focus3(event): 
    tpayee_field.focus_set() 
  
# set focus on the tcharity_field box 
def focus4(event): 
    tcharity_field.focus_set() 
  
# set focus on the thealthcare_field box 
def focus5(event): 
    thealthcare_field.focus_set() 
  
# set focus on the tauto_field box 
def focus6(event): 
    tauto_field.focus_set() 

# set focus on the tmedicql_field box  
def focus7(event): 
    tmedical_field.focus_set() 

# set focus on the ttaxes_field box 
def focus8(event): 
    ttaxes_field.focus_set() 

# set focus on the tother_field box
def focus9(event): 
    tother_field.focus_set() 

# Function for clearing the 
# contents of text entry boxes 
def clear(): 
    tdate_field.delete(0, END) 
    ttype_field.delete(0, END)  
    tpayee_field.delete(0, END) 
    tcharity_field.delete(0, END) 
    thealthcare_field.delete(0, END) 
    tauto_field.delete(0, END)
    tmedical_field.delete(0, END)
    ttaxes_field.delete(0, END)
    tother_field.delete(0, END)

# Function to take data from GUI  
# window and write to an spreadsheet file 
def insert():   
    # if user not fill any entry 
    # then print "empty input" 
    if (tdate_field.get() == "" and
        ttype_field.get() == "" and
        tpayee_field.get() == "" and
        tcharity_field.get() == "" and
        thealthcare_field.get() == "" and
        tauto_field.get() == "" and
        tmedical_field.get() == "" and
        ttaxes_field.get() == "" and
        tother_field.get() == ""):     
            print("empty input") 
  
    else: 
        # Get the first empty row to enter next record
        # Initially set active cell to first cell in Date column
        for cell in ws['A']:
            if cell.value is None:
                #print(cell.row) 
                current_row = cell.row
                break
        else:
            current_row = cell.row + 1
            #print(current_row)
        current_column = ws.max_column
  
        # get method returns current text 
        # as string which we write into 
        # the spreadsheet at particular location 
        ws.cell(row=current_row, column=1).value = tdate_field.get() 
        ws.cell(row=current_row, column=2).value = ttype_field.get()  
        ws.cell(row=current_row, column=3).value = tpayee_field.get() 
        ws.cell(row=current_row, column=4).value = tcharity_field.get() 
        ws.cell(row=current_row, column=5).value = thealthcare_field.get() 
        ws.cell(row=current_row, column=6).value = tauto_field.get()
        ws.cell(row=current_row, column=7).value = tmedical_field.get()
        ws.cell(row=current_row, column=8).value = ttaxes_field.get()
        ws.cell(row=current_row, column=9).value = tother_field.get()
  
        # save the file 
        wb.save('C:/Taxes/taxes-2021.xlsx') 
  
        # set focus on the tdate_field box 
        tdate_field.focus_set() 
  
        # call the clear() function 
        clear() 
  
  
# Driver code 
if __name__ == "__main__": 
      
    # create a GUI window 
    root = Tk()
  
    # set the background colour of GUI window 
    root.configure(background='light blue') 
  
    # set the title of GUI window 
    root.title("Tax Record Form") 
  
    # set the configuration of GUI window 
    root.geometry('450x450+200+200') 
  
    reset() 

    # create a Form label 
    heading = Label(root, text="Enter tax-related expenses:", anchor="c", bg="light blue", font=('Helvetica 24 bold italic')) 
  
    # create a Date label 
    tdate = Label(root, text="Date:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

    # create a Type label 
    ttype = Label(root, text="Type:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

    # create a Payee label 
    tpayee = Label(root, text="Payee:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

    # create a Charity lable 
    tcharity = Label(root, text="Charity:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

    # create a Healthcare label 
    thealthcare = Label(root, text="Healthcare:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

    # create a Auto label 
    tauto = Label(root, text="Auto:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

    # create a Medical label 
    tmedical = Label(root, text="Medical:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

    # create an Taxes label 
    ttaxes = Label(root, text="Taxes:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

    # create an Other label 
    tother = Label(root, text="Other:", bg="light blue", anchor="e", padx=3, pady=3, font=('Helvetica 18 bold'), width=10) 

  
    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    heading.grid(row=1, column=1, pady=6) 
    tdate.grid(row=3, column=0) 
    ttype.grid(row=5, column=0) 
    tpayee.grid(row=7, column=0) 
    tcharity.grid(row=9, column=0) 
    thealthcare.grid(row=11, column=0) 
    tauto.grid(row=13, column=0) 
    tmedical.grid(row=15, column=0) 
    ttaxes.grid(row=17, column=0) 
    tother.grid(row=19, column=0) 
  
    # create a text entry box 
    # for typing the information 
    tdate_field = Entry(root) 
    ttype_field = Entry(root) 
    tpayee_field = Entry(root) 
    tcharity_field = Entry(root) 
    thealthcare_field = Entry(root) 
    tauto_field = Entry(root) 
    tmedical_field = Entry(root) 
    ttaxes_field = Entry(root) 
    tother_field = Entry(root) 

  
    # bind method of widget is used for 
    # the binding the function with the events 
    # whenever the enter key is pressed 
    # then call the focus1 function 
    tdate_field.bind("<Return>", focus1) 
  
    # whenever the enter key is pressed 
    # then call the focus2 function 
    ttype_field.bind("<Return>", focus2) 

    # whenever the enter key is pressed 
    # then call the focus4 function 
    tpayee_field.bind("<Return>", focus3) 
  
    # whenever the enter key is pressed 
    # then call the focus5 function 
    tcharity_field.bind("<Return>", focus4) 
  
    # whenever the enter key is pressed 
    # then call the focus6 function 
    thealthcare_field.bind("<Return>", focus5) 
  
    # whenever the enter key is pressed 
    # then call the focus6 function 
    tauto_field.bind("<Return>", focus6)

    # whenever the enter key is pressed 
    # then call the focus6 function 
    tmedical_field.bind("<Return>", focus7)

    # whenever the enter key is pressed 
    # then call the focus6 function 
    ttaxes_field.bind("<Return>", focus8) 

    # whenever the enter key is pressed 
    # then call the focus6 function 
    tother_field.bind("<Return>", focus9) 


    # grid method is used for placing 
    # the widgets at respective positions 
    # in table like structure . 
    tdate_field.grid(row=3, column=1, ipadx="20") 
    ttype_field.grid(row=5, column=1, ipadx="10") 
    tpayee_field.grid(row=7, column=1, ipadx="10") 
    tcharity_field.grid(row=9, column=1, ipadx="10") 
    thealthcare_field.grid(row=11, column=1, ipadx="10") 
    tauto_field.grid(row=13, column=1, ipadx="10") 
    tmedical_field.grid(row=15, column=1, ipadx="10") 
    ttaxes_field.grid(row=17, column=1, ipadx="10") 
    tother_field.grid(row=19, column=1, ipadx="10") 

    # call reset function 
    reset()

    # Place focus on first input field, Date
    tdate_field.focus_set()
    
    # create a Submit Button and place into the root window 
    submit = Button(root, text="Submit", fg="Black", bd=2,
                            bg="Red", pady=10, command=insert) 
    submit.grid(row=25, column=1) 

    
    # start the GUI 
    root.mainloop() 